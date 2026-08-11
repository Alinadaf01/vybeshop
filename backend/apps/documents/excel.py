"""Shared Excel-export styling per BACKEND-TASK.md §6 — used by every admin
`.xlsx` download so they look like one consistent product, not ad-hoc dumps.

openpyxl has no equivalent of the PDF pipeline's `@font-face` embedding —
Excel resolves font *names* against whatever is installed on the machine
that opens the file, so trying to force "Peyda" here would either silently
fall back to a default font (harmless) or, on a machine that happens to
have a font of that name for unrelated reasons, render wrong. Font choice
is left to Excel's own default; every other part of the §6 spec (RTL sheet,
graphite header row, frozen header, autofilter, number formats, info row,
totals row) is enforced here.
"""

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .persian import format_jalali_date

GRAPHITE = "0B0B0C"
TOMAN_FORMAT = '#,##0" تومان"'
COUNT_FORMAT = "#,##0"

_HEADER_FILL = PatternFill(fill_type="solid", fgColor=GRAPHITE)
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_INFO_FONT = Font(bold=True, size=12)
_META_FONT = Font(italic=True, color="7A7D82", size=9)
_TOTALS_FONT = Font(bold=True)


class Column:
    """A workbook column. `number_format` is `None` for plain text,
    `COUNT_FORMAT` for a thousands-grouped integer (e.g. quantities), or
    `TOMAN_FORMAT` for a currency amount — the §6 requirement is specific
    that only money gets the "تومان" suffix, not every number."""

    __slots__ = ("field", "header", "number_format")

    def __init__(self, field: str, header: str, number_format: str | None = None):
        self.field = field
        self.header = header
        self.number_format = number_format


def build_workbook(
    *,
    sheet_name: str,
    report_title: str,
    columns: list[Column],
    rows: list[dict],
    filter_summary: str = "",
    generated_by: str = "",
    totals: dict | None = None,
) -> Workbook:
    """`rows` are plain dicts keyed by each column's `field`. `totals`, if
    given, is `{field: value}` for a bold summary row appended under the
    data — only numeric reports need one (§6: "ردیف جمع کل برای گزارش‌های
    عددی")."""
    workbook = Workbook()
    sheet: Worksheet = workbook.active
    sheet.title = sheet_name[:31]
    sheet.sheet_view.rightToLeft = True

    sheet.append([report_title])
    sheet["A1"].font = _INFO_FONT

    info_bits = [filter_summary] if filter_summary else []
    if generated_by:
        info_bits.append(f"تهیه‌شده توسط {generated_by}")
    info_bits.append(f"تاریخ تولید: {format_jalali_date(timezone.localtime(timezone.now()))}")
    sheet.append([" | ".join(info_bits)])
    sheet["A2"].font = _META_FONT
    sheet.append([])

    header_row = 4
    sheet.append([column.header for column in columns])
    for col_idx in range(1, len(columns) + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    sheet.row_dimensions[header_row].height = 24

    data_start = header_row + 1
    for row in rows:
        sheet.append([row.get(column.field) for column in columns])
        r = sheet.max_row
        for col_idx, column in enumerate(columns, start=1):
            if column.number_format:
                cell = sheet.cell(row=r, column=col_idx)
                cell.number_format = column.number_format
                cell.alignment = Alignment(horizontal="right")

    last_data_row = sheet.max_row

    if totals:
        sheet.append([totals.get(column.field, "") for column in columns])
        r = sheet.max_row
        if not sheet.cell(row=r, column=1).value:
            sheet.cell(row=r, column=1, value="جمع کل")
        for col_idx, column in enumerate(columns, start=1):
            cell = sheet.cell(row=r, column=col_idx)
            cell.font = _TOTALS_FONT
            if column.number_format:
                cell.number_format = column.number_format
                cell.alignment = Alignment(horizontal="right")

    last_col_letter = get_column_letter(len(columns))
    sheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_data_row}"
    sheet.freeze_panes = f"A{data_start}"

    for col_idx, column in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        lengths = [len(str(column.header))] + [len(str(row.get(column.field, ""))) for row in rows]
        if totals:
            lengths.append(len(str(totals.get(column.field, ""))))
        sheet.column_dimensions[letter].width = min(max(max(lengths) + 2, 10), 45)

    return workbook
